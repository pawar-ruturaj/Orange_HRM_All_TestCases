# we can create test case data by using Excel.
# we create a new XLutils.py file in utilities, and in that file we create method about exel.
# now we create new test case file test_addemp_ddt file for data-driven run to add employee test case.
# we create another method in XLutils.py file for ditect rows and columns of that excel file sheet.

import openpyxl

# this method get rows from Excel.
def getrowCount(file, sheetname):
    book = openpyxl.load_workbook(file)  # which Excel file we need to load. give Excel file path in test case.
    sheet = book[sheetname]  # here we specify sheetname which is in Excel file.
    return (sheet.max_row)  # it will return maximum rows in that sheet.


# this method for read data from excel
def readData(file, sheetname, rowno, colno):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=rowno, column=colno).value


# this method for write data in excel
def writeData(file, sheetname, rowno, colno, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=rowno, column=colno).value = data
    book.save(file)
